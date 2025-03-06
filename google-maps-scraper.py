import time
import random
import datetime
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.styles import Font

def iniciar_driver():
    """Inicia o WebDriver com configurações otimizadas."""
    options = Options()
    options.add_argument("--headless")  # Executar sem interface gráfica
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920x1080")
    options.add_argument("--log-level=3")  # Reduz logs desnecessários
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64)")

    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)

def buscar_empresas_google_maps(query, cidade, num_paginas=2):
    """Coleta empresas do Google Maps e evita duplicatas."""
    driver = iniciar_driver()
    url = f"https://www.google.com/maps/search/{query}+{cidade}"
    driver.get(url)
    time.sleep(random.uniform(4, 6))  # Espera inicial com tempo aleatório

    empresas = []
    nomes_existentes = set()  # Conjunto para evitar duplicatas

    for _ in range(num_paginas):
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, "Nv2PK"))
            )

            lista_empresas = driver.find_elements(By.CLASS_NAME, "Nv2PK")

            for empresa in lista_empresas[:10]:  
                try:
                    empresa.click()
                    time.sleep(random.uniform(3, 5))  

                    try:
                        nome = WebDriverWait(driver, 5).until(
                            EC.presence_of_element_located((By.CLASS_NAME, "DUwDvf"))
                        ).text
                    except:
                        nome = "N/A"

                    try:
                        endereco = driver.find_element(By.CLASS_NAME, "CsEnBe").text
                    except:
                        endereco = "N/A"

                    try:
                        telefone = driver.find_element(By.CSS_SELECTOR, "button[data-item-id='phone']").text
                    except:
                        telefone = "N/A"

                    if nome not in nomes_existentes:  # Evita duplicatas
                        empresas.append({"Nome": nome, "Telefone": telefone, "Endereço": endereco})
                        nomes_existentes.add(nome)

                    driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
                    time.sleep(random.uniform(2, 3))

                except Exception as e:
                    print(f"Erro ao processar empresa: {e}")
                    continue

            driver.find_element(By.TAG_NAME, "body").send_keys(Keys.PAGE_DOWN)
            time.sleep(random.uniform(3, 5))

        except Exception as e:
            print(f"Ocorreu um erro: {e}")
            break

    driver.quit()
    return empresas

def salvar_dados_excel(dados, query, cidade):
    """Salva os dados em um arquivo Excel formatado."""
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    nome_arquivo = f"{query}_{cidade}_{timestamp}.xlsx".replace(" ", "_")

    wb = Workbook()
    ws = wb.active
    ws.title = "Empresas"
    ws.append(["Nome", "Telefone", "Endereço"])

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for empresa in dados:
        ws.append([empresa["Nome"], empresa["Telefone"], empresa["Endereço"]])
        ws[f"A{ws.max_row}"].font = Font(bold=True)
        ws[f"C{ws.max_row}"].font = Font(italic=True)

    # Ajusta a largura das colunas automaticamente
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(nome_arquivo)
    print(f"✅ Dados salvos em {nome_arquivo}")

if __name__ == "__main__":
    query = input("Digite o tipo de estabelecimento: ")  
    cidade = input("Digite a cidade: ")  

    resultados = buscar_empresas_google_maps(query, cidade)
    
    if resultados:
        salvar_dados_excel(resultados, query, cidade)
    else:
        print("❌ Nenhum dado foi extraído.")
